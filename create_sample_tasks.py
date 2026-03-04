"""
Script to generate sample development tasks Excel file.
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_tasks_file(filepath: str = "tasks/development_tasks.xlsx"):
    """Create a sample development tasks Excel file."""
    
    # Ensure directory exists
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Development Tasks"
    
    # Headers
    headers = [
        "Task ID", "Title", "Description", "Type", "Target Component",
        "Priority", "Status", "Branch Name", "PR URL", "Created Date", "Updated Date"
    ]
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Sample tasks
    sample_tasks = [
        {
            "Task ID": "TASK-001",
            "Title": "Add password reset feature",
            "Description": "Implement forgot password flow with email verification. User should be able to request password reset, receive email with token, and set new password.",
            "Type": "feature",
            "Target Component": "backend",
            "Priority": "high",
            "Status": "new",
            "Branch Name": "",
            "PR URL": "",
            "Created Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Updated Date": ""
        },
        {
            "Task ID": "TASK-002",
            "Title": "Add user profile page",
            "Description": "Create a user profile page in the frontend where users can view and edit their profile information including avatar upload.",
            "Type": "feature",
            "Target Component": "frontend",
            "Priority": "medium",
            "Status": "new",
            "Branch Name": "",
            "PR URL": "",
            "Created Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Updated Date": ""
        },
        {
            "Task ID": "TASK-003",
            "Title": "Add API rate limiting",
            "Description": "Implement rate limiting on the API to prevent abuse. Should limit requests per IP/user with configurable thresholds.",
            "Type": "enhancement",
            "Target Component": "backend",
            "Priority": "medium",
            "Status": "new",
            "Branch Name": "",
            "PR URL": "",
            "Created Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Updated Date": ""
        },
        {
            "Task ID": "TASK-004",
            "Title": "Fix login session timeout",
            "Description": "Users are being logged out after 5 minutes. The session should persist for 24 hours.",
            "Type": "bugfix",
            "Target Component": "backend",
            "Priority": "high",
            "Status": "new",
            "Branch Name": "",
            "PR URL": "",
            "Created Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Updated Date": ""
        }
    ]
    
    # Write tasks
    for row, task in enumerate(sample_tasks, 2):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=task.get(header, ""))
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Adjust column widths
    column_widths = [12, 30, 60, 12, 18, 10, 12, 30, 40, 20, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    
    # Save
    wb.save(filepath)
    print(f"✅ Created tasks file: {filepath}")
    print(f"   Added {len(sample_tasks)} sample tasks")
    return filepath


if __name__ == "__main__":
    create_tasks_file()
