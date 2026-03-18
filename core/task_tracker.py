"""
Task Tracker Module - Manages development tasks from Excel/JSON.

Supports:
- Reading tasks from Excel (.xlsx) or JSON
- Writing task status updates
- Auto-detection of new tasks
- Task prioritization
"""

import os
import json
from datetime import datetime
from typing import Any, Dict, List, Optional
from dataclasses import dataclass, field, asdict
from enum import Enum
from pathlib import Path


class TaskStatus(Enum):
    """Status of a development task."""
    NEW = "new"
    IN_PROGRESS = "in_progress"
    DEVELOPMENT_COMPLETE = "development_complete"
    PR_CREATED = "pr_created"
    PR_REVIEW = "pr_review"
    PR_APPROVED = "pr_approved"
    MERGED = "merged"
    COMPLETED = "completed"
    FAILED = "failed"


class TaskPriority(Enum):
    """Priority of a development task."""
    LOW = "low"
    MEDIUM = "medium"
    HIGH = "high"
    CRITICAL = "critical"


@dataclass
class DevelopmentTask:
    """Represents a development task/enhancement."""
    id: str
    title: str
    description: str
    type: str  # feature, bugfix, enhancement, refactor
    priority: str = "medium"
    status: str = "new"
    assigned_to: Optional[str] = None
    target_component: str = "backend"  # backend, frontend, both
    created_at: str = field(default_factory=lambda: datetime.now().isoformat())
    started_at: Optional[str] = None
    completed_at: Optional[str] = None
    pr_url: Optional[str] = None
    pr_number: Optional[int] = None
    branch_name: Optional[str] = None
    files_changed: List[str] = field(default_factory=list)
    error_message: Optional[str] = None
    metadata: Dict[str, Any] = field(default_factory=dict)
    
    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)
    
    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "DevelopmentTask":
        clean_data = {k: v for k, v in data.items() if k in cls.__dataclass_fields__}

        task_id = str(clean_data.get("id") or "").strip()
        if not task_id:
            task_id = f"TASK-{datetime.now().strftime('%Y%m%d%H%M%S')}"

        title = str(clean_data.get("title") or "").strip() or f"Task {task_id}"
        description = str(clean_data.get("description") or "").strip() or title
        task_type = str(clean_data.get("type") or "").strip() or "enhancement"

        clean_data["id"] = task_id
        clean_data["title"] = title
        clean_data["description"] = description
        clean_data["type"] = task_type

        return cls(**clean_data)


class TaskTracker:
    """
    Manages development tasks from Excel or JSON files.
    
    Features:
    - Load tasks from Excel (.xlsx) or JSON
    - Track task status changes
    - Detect new tasks automatically
    - Persist task updates
    """
    
    def __init__(self, tasks_file: str = "tasks/development_tasks.xlsx"):
        """
        Initialize task tracker.
        
        Args:
            tasks_file: Path to tasks file (Excel or JSON)
        """
        self.tasks_file = Path(tasks_file)
        self.tasks: Dict[str, DevelopmentTask] = {}
        self._last_modified: Optional[float] = None
        self._ensure_tasks_dir()
        self._load_tasks()
    
    def _ensure_tasks_dir(self) -> None:
        """Ensure tasks directory exists."""
        self.tasks_file.parent.mkdir(parents=True, exist_ok=True)
    
    def _load_tasks(self) -> None:
        """Load tasks from file."""
        if not self.tasks_file.exists():
            self._create_default_tasks_file()
            return
        
        if self.tasks_file.suffix == ".xlsx":
            self._load_from_excel()
        elif self.tasks_file.suffix == ".json":
            self._load_from_json()
        else:
            raise ValueError(f"Unsupported file type: {self.tasks_file.suffix}")
        
        self._last_modified = self.tasks_file.stat().st_mtime
    
    def _load_from_excel(self) -> None:
        """Load tasks from Excel file."""
        try:
            import openpyxl
        except ImportError:
            print("Installing openpyxl for Excel support...")
            import subprocess
            subprocess.run(["pip", "install", "openpyxl", "-q"])
            import openpyxl
        
        wb = openpyxl.load_workbook(self.tasks_file)
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value.lower().replace(" ", "_") if cell.value else f"col_{i}" 
                   for i, cell in enumerate(ws[1])]
        
        # Map Excel columns to DevelopmentTask fields
        field_mapping = {
            "task_id": "id",
            "id": "id",
            "title": "title",
            "name": "title",
            "description": "description",
            "desc": "description",
            "type": "type",
            "task_type": "type",
            "priority": "priority",
            "status": "status",
            "component": "target_component",
            "target_component": "target_component",
            "assigned_to": "assigned_to",
            "assignee": "assigned_to",
            "pr_url": "pr_url",
            "pr_number": "pr_number",
            "branch": "branch_name",
            "branch_name": "branch_name",
        }
        
        # Read data rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            
            row_data = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    header = headers[i]
                    mapped_field = field_mapping.get(header, header)
                    if mapped_field in DevelopmentTask.__dataclass_fields__:
                        row_data[mapped_field] = value
            
            # Ensure required fields
            if "id" not in row_data:
                continue
            if "title" not in row_data:
                row_data["title"] = f"Task {row_data['id']}"
            if "description" not in row_data:
                row_data["description"] = row_data.get("title", "")
            if "type" not in row_data:
                row_data["type"] = "enhancement"
            
            # Clean up None values
            row_data = {k: v for k, v in row_data.items() if v is not None}
            
            try:
                task = DevelopmentTask.from_dict(row_data)
                self.tasks[task.id] = task
            except Exception:
                continue
        
        wb.close()
    
    def _load_from_json(self) -> None:
        """Load tasks from JSON file."""
        with open(self.tasks_file, "r") as f:
            data = json.load(f)
        
        tasks_list = data.get("tasks", data) if isinstance(data, dict) else data
        
        for task_data in tasks_list:
            try:
                task = DevelopmentTask.from_dict(task_data)
                self.tasks[task.id] = task
            except Exception:
                continue
    
    def _create_default_tasks_file(self) -> None:
        """Create a default tasks file."""
        if self.tasks_file.suffix == ".xlsx":
            self._create_excel_template()
        else:
            self._create_json_template()
    
    def _create_excel_template(self) -> None:
        """Create Excel template for tasks."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            import subprocess
            subprocess.run(["pip", "install", "openpyxl", "-q"])
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Development Tasks"
        
        # Headers
        headers = [
            "Task ID", "Title", "Description", "Type", "Priority", 
            "Status", "Component", "Assigned To", "PR URL", "PR Number", "Branch"
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Sample task
        sample_task = [
            "TASK-001", 
            "Add user authentication", 
            "Implement JWT-based authentication for API endpoints",
            "feature",
            "high",
            "new",
            "backend",
            "",
            "",
            "",
            ""
        ]
        for col, value in enumerate(sample_task, 1):
            ws.cell(row=2, column=col, value=value)
        
        # Adjust column widths
        column_widths = [12, 30, 50, 12, 10, 15, 12, 15, 40, 12, 25]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        wb.save(self.tasks_file)
        wb.close()
        print(f"Created Excel template: {self.tasks_file}")
    
    def _create_json_template(self) -> None:
        """Create JSON template for tasks."""
        template = {
            "tasks": [
                {
                    "id": "TASK-001",
                    "title": "Add user authentication",
                    "description": "Implement JWT-based authentication for API endpoints",
                    "type": "feature",
                    "priority": "high",
                    "status": "new",
                    "target_component": "backend"
                }
            ]
        }
        with open(self.tasks_file, "w") as f:
            json.dump(template, f, indent=2)
        print(f"Created JSON template: {self.tasks_file}")
    
    def save_tasks(self) -> None:
        """Save tasks back to file."""
        if self.tasks_file.suffix == ".xlsx":
            self._save_to_excel()
        else:
            self._save_to_json()
    
    def _save_to_excel(self) -> None:
        """Save tasks to Excel file."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Development Tasks"
        
        # Headers
        headers = [
            "Task ID", "Title", "Description", "Type", "Priority", 
            "Status", "Component", "Assigned To", "PR URL", "PR Number", 
            "Branch", "Created At", "Started At", "Completed At"
        ]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Write tasks
        for row, task in enumerate(self.tasks.values(), 2):
            ws.cell(row=row, column=1, value=task.id)
            ws.cell(row=row, column=2, value=task.title)
            ws.cell(row=row, column=3, value=task.description)
            ws.cell(row=row, column=4, value=task.type)
            ws.cell(row=row, column=5, value=task.priority)
            ws.cell(row=row, column=6, value=task.status)
            ws.cell(row=row, column=7, value=task.target_component)
            ws.cell(row=row, column=8, value=task.assigned_to)
            ws.cell(row=row, column=9, value=task.pr_url)
            ws.cell(row=row, column=10, value=task.pr_number)
            ws.cell(row=row, column=11, value=task.branch_name)
            ws.cell(row=row, column=12, value=task.created_at)
            ws.cell(row=row, column=13, value=task.started_at)
            ws.cell(row=row, column=14, value=task.completed_at)
            
            # Color code by status
            status_colors = {
                "new": "E2EFDA",
                "in_progress": "FFF2CC",
                "development_complete": "DDEBF7",
                "pr_created": "FCE4D6",
                "pr_review": "E4DFEC",
                "pr_approved": "D9EAF7",
                "merged": "C6EFCE",
                "completed": "C6EFCE",
                "failed": "FFC7CE"
            }
            fill_color = status_colors.get(task.status, "FFFFFF")
            for col in range(1, 15):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid"
                )
        
        wb.save(self.tasks_file)
        wb.close()
    
    def _save_to_json(self) -> None:
        """Save tasks to JSON file."""
        with open(self.tasks_file, "w") as f:
            json.dump({
                "tasks": [task.to_dict() for task in self.tasks.values()],
                "last_updated": datetime.now().isoformat()
            }, f, indent=2)
    
    def get_new_tasks(self) -> List[DevelopmentTask]:
        """Get all tasks with 'new' status."""
        return [t for t in self.tasks.values() if t.status == TaskStatus.NEW.value]
    
    def get_pending_tasks(self) -> List[DevelopmentTask]:
        """Get tasks that need processing."""
        pending_statuses = [TaskStatus.NEW.value, TaskStatus.IN_PROGRESS.value]
        return [t for t in self.tasks.values() if t.status in pending_statuses]
    
    def get_task(self, task_id: str) -> Optional[DevelopmentTask]:
        """Get a specific task by ID."""
        return self.tasks.get(task_id)
    
    def update_task_status(
        self, 
        task_id: str, 
        status: TaskStatus,
        **kwargs
    ) -> Optional[DevelopmentTask]:
        """Update task status and optional fields."""
        task = self.tasks.get(task_id)
        if not task:
            return None
        
        task.status = status.value
        
        # Update timestamps
        if status == TaskStatus.IN_PROGRESS and not task.started_at:
            task.started_at = datetime.now().isoformat()
        elif status in [TaskStatus.COMPLETED, TaskStatus.MERGED]:
            task.completed_at = datetime.now().isoformat()
        
        # Update additional fields
        for key, value in kwargs.items():
            if hasattr(task, key):
                setattr(task, key, value)
        
        self.save_tasks()
        return task
    
    def has_new_tasks(self) -> bool:
        """Check if file was modified and has new tasks."""
        if not self.tasks_file.exists():
            return False
        
        current_mtime = self.tasks_file.stat().st_mtime
        if self._last_modified and current_mtime > self._last_modified:
            self._load_tasks()
            return len(self.get_new_tasks()) > 0
        
        return len(self.get_new_tasks()) > 0
    
    def add_task(self, task: DevelopmentTask) -> None:
        """Add a new task."""
        self.tasks[task.id] = task
        self.save_tasks()
    
    def get_tasks_by_status(self, status: TaskStatus) -> List[DevelopmentTask]:
        """Get all tasks with a specific status."""
        return [t for t in self.tasks.values() if t.status == status.value]
    
    def get_all_tasks(self) -> List[DevelopmentTask]:
        """Get all tasks."""
        return list(self.tasks.values())


if __name__ == "__main__":
    # Test the tracker
    tracker = TaskTracker("tasks/development_tasks.xlsx")
    print(f"Loaded {len(tracker.tasks)} tasks")
    for task in tracker.get_all_tasks():
        print(f"  - {task.id}: {task.title} [{task.status}]")
