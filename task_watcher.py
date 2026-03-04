"""
Task Watcher - Monitors Excel file for new development tasks
and auto-triggers the SDLC agent workflow
"""

import time
import os
from pathlib import Path
from datetime import datetime
import openpyxl
from rich.console import Console
from rich.panel import Panel
from rich.progress import Progress, SpinnerColumn, TextColumn

console = Console()

class TaskWatcher:
    """Watches for new tasks in Excel and triggers workflow."""
    
    def __init__(self, tasks_file='tasks/development_tasks.xlsx'):
        self.tasks_file = tasks_file
        self.processed_tasks = set()
        self.check_interval = 5  # Check every 5 seconds
        
    def get_tasks(self):
        """Load tasks from Excel file."""
        if not os.path.exists(self.tasks_file):
            return []
        
        try:
            wb = openpyxl.load_workbook(self.tasks_file)
            ws = wb.active
            
            tasks = []
            for row in ws.iter_rows(min_row=2, values_only=False):
                if row[0].value:  # Task ID
                    task = {
                        'id': row[0].value,
                        'title': row[1].value,
                        'description': row[2].value,
                        'status': row[3].value,
                        'assigned_to': row[4].value if len(row) > 4 else None,
                        'created_date': row[5].value if len(row) > 5 else None,
                    }
                    tasks.append(task)
            
            return tasks
        except Exception as e:
            console.print(f"[red]Error reading tasks: {e}[/red]")
            return []
    
    def update_task_status(self, task_id, new_status):
        """Update task status in Excel."""
        try:
            wb = openpyxl.load_workbook(self.tasks_file)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2):
                if row[0].value == task_id:
                    row[3].value = new_status
                    row[4].value = datetime.now().isoformat()  # Updated timestamp
                    break
            
            wb.save(self.tasks_file)
        except Exception as e:
            console.print(f"[red]Error updating task: {e}[/red]")
    
    def watch(self):
        """Watch for new tasks and trigger workflow."""
        console.print("\n[bold cyan]🔍 Task Watcher Started[/bold cyan]")
        console.print(f"[dim]Monitoring: {self.tasks_file}[/dim]\n")
        
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            console=console,
            transient=True,
        ) as progress:
            task = progress.add_task("[cyan]Watching for new tasks...", total=None)
            
            while True:
                try:
                    new_tasks = self.get_tasks()
                    
                    for task_obj in new_tasks:
                        if task_obj['status'] == 'pending' and task_obj['id'] not in self.processed_tasks:
                            console.print(Panel(
                                f"[bold green]✓ New Task Detected![/bold green]\n\n"
                                f"[cyan]ID:[/cyan] {task_obj['id']}\n"
                                f"[cyan]Title:[/cyan] {task_obj['title']}\n"
                                f"[cyan]Description:[/cyan] {task_obj['description']}",
                                title="[bold yellow]Task Detected[/bold yellow]",
                                border_style="green"
                            ))
                            
                            # Mark as processing
                            self.update_task_status(task_obj['id'], 'processing')
                            
                            # Trigger workflow
                            self.trigger_workflow(task_obj)
                            
                            self.processed_tasks.add(task_obj['id'])
                    
                    # Update progress
                    progress.update(task, description=f"[cyan]✓ Monitoring ({len(self.processed_tasks)} tasks processed)")
                    time.sleep(self.check_interval)
                    
                except KeyboardInterrupt:
                    console.print("\n[yellow]Watcher stopped by user[/yellow]")
                    break
                except Exception as e:
                    console.print(f"[red]Error: {e}[/red]")
                    time.sleep(self.check_interval)
    
    def trigger_workflow(self, task):
        """Trigger the SDLC agent workflow for the task."""
        import subprocess
        import sys
        
        console.print(f"\n[bold yellow]⚙️  Triggering SDLC Workflow for {task['id']}...[/bold yellow]")
        
        try:
            # Run the main orchestrator
            cmd = [
                sys.executable,
                'run_sdlc.py',
                '--task', task['id'],
                '--goal', task['description']
            ]
            
            result = subprocess.run(cmd, capture_output=True, text=True)
            
            if result.returncode == 0:
                console.print(f"[green]✓ Workflow completed for {task['id']}[/green]")
                self.update_task_status(task['id'], 'completed')
            else:
                console.print(f"[red]✗ Workflow failed: {result.stderr}[/red]")
                self.update_task_status(task['id'], 'failed')
                
        except Exception as e:
            console.print(f"[red]Error triggering workflow: {e}[/red]")
            self.update_task_status(task['id'], 'failed')


if __name__ == '__main__':
    watcher = TaskWatcher()
    watcher.watch()
